"""
版权所有 (c) 2025 [吴宪]

本软件TermPlus（以下简称“软件”）为个人非商业用途开源发布。任何人均可免费使用、复制、修改和分发本软件，但仅限于个人非商业用途。
对于商业用途（包括但不限于嵌入商业产品、提供商业服务、营利分发等），必须事先获得作者的书面许可。

本软件按“现状”提供，不附带任何明示或暗示的担保，包括但不限于对适销性、特定用途适用性以及不侵权的保证。
在任何情况下，作者均不对因使用或无法使用本软件而产生的任何直接、间接、偶然、特殊或后续损害承担责任。

如需商业授权或有任何疑问，请联系：[dakongwuxian@gmail.com]
"""
# 标准库导入
import base64
import os
import re
import tempfile
import threading
import webbrowser
from io import BytesIO
import tkinter as tk
from tkinter import ttk, filedialog, messagebox

# 第三方库导入
from PIL import Image, ImageTk

# 可选的第三方库
try:
    import openpyxl
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

try:
    from tkinterdnd2 import DND_FILES, TkinterDnD
    TKINTERDND2_AVAILABLE = True
except ImportError:
    TKINTERDND2_AVAILABLE = False 

# 嵌入的二维码 base64 字符串（PNG 格式）
img_base64 = (
"iVBORw0KGgoAAAANSUhEUgAAANgAAADZAQAAAABqB8PlAAAAGmVYSWZNTQAqAAAACAABARIAAwAA"
"AAEAAQAAAAAAABPAdecAAATCSURBVHjazVjbjuNEED3ltmgLJDp8wKS9/+EdZzV8CJ+wvGWkRenR"
"rJR5Wz6FD1iYzmSl5SMQsmd546WNBOqA7cNDkrlkLwybGOG3uNLlrqpT51S3EO99Evy3NgQARVnT"
"hKjPPaQyJsxiAcAOs5cZW3V5jDDS7hnK7rhpjGQo2Q8U+0ocrphLkPgjFsmbipg5dJIjAeC3f3wJ"
"AKMD5bqsK4yeXnQTYWdFMuV31l3c/kwPWdt+HcPk1nOXdEmwxq1A+6YHgG44nGl6MLeoV07RS49m"
"dOagWA3yPQkjoMC1BETXTjCGAS5S/+oGu7ur2/16hSS56EFgquilN8bMSZIcJp8e+AlI3HdIZzgr"
"u2/TP7QDPCrQAcqPrQnQDqoSBkxRetvv0bfuPpbWPDXZ8MSMHpYBdMpjDMOg2ar6sLFnN7bpSk48"
"AGS6haoZREe89keSD8KDdFAeAIOmU156Q8wcynqo2k716mRBjoAzfv/EX6+J4/lRsrsuyQ+rARk8"
"+p3Yp+WmtlF5jAFCz71if2ANaDe8pLE46RUMRa9SVQPCi9VE1bkdprZ0Jb3QBLJVHpaEdkVhh6kt"
"YpyikB4E2UJIGzBVy2Xd2zRe3HfDsLcGIGieL1kZBmB2vqxsMJyzBSyTTQOP7rTyobAUAMD73XlC"
"t0XZA8AUUKwZwoz0Uj20tjIn2/fYPtnliRSz7LU/EY7COsRRY4Nkv0+sVHfWjf4xvuzBuZ57lBXC"
"Wg0YDQOmherHA2E3ovSqZjB2/T3hC9KXvd1g924r84bBD7CXu2LgN3uZAlJXwaxj74EwBTA+4Kwo"
"93Idyxa2Ctw8IRiHsh2obxPjLrqJXGHUrCmXdtREdJN6GM1JACh/9PjYmPWLvwLxIipvt7Nwfi8d"
"+2NpHoFjABss9RaYTgH7rt4MB4jPbfHptvgEUiAHkmaVqeVzocFXAIA/fzPM1HLphzrnaFeUi8pu"
"+7Y3IUCzfTgv/ctc6znbkhXClidM0BGKNXfmrA/7jA/VHB0BVVkas+0VzF7HouxtEm944o43n6La"
"9flqkTw49uztl65FAiDFU5SeuTRYneovcZT0QnQni+vrj69ttqtet4KWaJl/8/hJvQCgU5RXRwLJ"
"tAPzQbCUNhEn885eh4DVZ233qDJmNP1aVEsMNJvOWkhvA01EyWq8ns/K+h524wE1Z6q4qCsTNM9Z"
"2R7ALELxrV6pDsODMb2j0nkuuFW2gFlbLnrQ6KiWJGkwhfIy0OzGOZckKwPSFaoXGsMIVQ1zH5IC"
"6B6ryxxmlal2kY+b2q5cgbIqB8KSpsf4c8N5LJSXygQz53J5p7ZfAG632fbRgD6Da/ESeNnjDPgU"
"z371P/hNbQupqwA9d1A9DAMASM19eOKD5/dWKpBxds4lKxjoeJB54p063YjI5Hg8bkRneNXJI4IO"
"3eS6HuYMG2Zsi7KXEABFLwwh4v68xHdIy8djdyXPwBy21vFMpBODBlik46S/d7e242a/2H9+25Bv"
"bb+8qZpRc9qq1aMKAdqV9ONDcXKy27eA9HUwWeyeP7k6aiRc0L8qFvkwPKHZTY6v7ahuXLoErgFj"
"vhZIN9Bd3pmk6tLmAVitTk+Px8ZIRFFeJu9fl+9xl/c/uhP/kO1vBYeg5BFXBCUAAAAASUVORK5C"
"YII="
)

class MainGUI(tk.Tk):
    def __init__(self):
        # 尝试使用 tkinterdnd2 创建窗口
        try:
            from tkinterdnd2 import TkinterDnD
            # 如果有 tkinterdnd2，使用它初始化
            TkinterDnD.Tk.__init__(self)
        except ImportError:
            # 如果没有，使用普通 tk.Tk
            super().__init__()

        self.title("Text Catcher - Regex Tester")
        self.geometry("1200x800")

        style = ttk.Style()
        # 定义一个名为 'Border.TFrame' 的新样式
        style.configure('Border.TFrame', 
                        relief="solid",         # 设置扁平边框，也可以是 'raised' 等
                        borderwidth=1,          # 设置边框宽度
                        background="lightgray"  # 确保背景颜色与边框区分开
                       )

        # 创建菜单栏
        self.menu_bar = tk.Menu(self)
        self.config(menu=self.menu_bar)

        # About 菜单
        self.about_menu = tk.Menu(self.menu_bar, tearoff=0)
        self.menu_bar.add_cascade(label="About", menu=self.about_menu)
        self.about_menu.add_command(label="Developed by Xian.Wu", state="disabled")
        self.about_menu.add_command(label="dakongwuxian@gmail.com", state="disabled")

        self.about_menu.add_command(label="Ver. 20260112", state="disabled")

        self.about_menu.add_command(label="Buy me a coffee ☕",command=self.show_about_window,state="normal")


        # 用于处理input text中的文本的信息
        self.input_string_content = None
        self.highlight_items = []
        
        # 查找功能相关
        self.search_history = []
        self.find_window = None
        
        # 位置映射表：{(row, col): (start_pos, end_pos)}
        self.match_positions = {}


        # 主框架
        self.main_frame = ttk.Frame(self, padding=10)
        self.main_frame.pack(fill="both", expand=True)

        # 2列均分
        self.main_frame.columnconfigure(0, weight=1, uniform="col")
        self.main_frame.columnconfigure(1, weight=1, uniform="col")
        #main_frame.columnconfigure(2, weight=1, uniform="col")

        self.main_frame.rowconfigure(0, weight=0)
        #self.main_frame.rowconfigure(0, minsize=100)
        self.main_frame.rowconfigure(1, weight=1)

        # 上方frame ————————————————————————————————————————————————————————————————————————————————————

        self.up_frame = ttk.Frame(self.main_frame) #, style='Border.TFrame'
        self.up_frame.grid(row=0, column=0,sticky="nsew", padx=5,columnspan=2)
        # 让 up_frame 的列的权重为1, 才能实现后面的sticky ew的自动占满
        self.up_frame.columnconfigure(0, weight=1)

        self.up_frame.rowconfigure(0,weight=0)
        self.up_frame.rowconfigure(1,weight=0)

        # self.up_frame.rowconfigure(1,weight=0)
        # self.up_frame.rowconfigure(2,weight=1)
        # self.up_frame.rowconfigure(3,weight=0)

        self.regular_expression_area_label = ttk.Label(self.up_frame, text="Regular Expression",font=("微软雅黑", 11,"bold"))
        self.regular_expression_area_label.grid(row=0,column=0,sticky="w", pady=5)

        self.up_button_frame = ttk.Frame(self.up_frame) #, style='Border.TFrame'
        self.up_button_frame.grid(row=1, column=0,sticky="ew", padx=5)
        self.up_button_frame.columnconfigure(4, weight=1)


        self.load_regex_button = ttk.Button(self.up_button_frame,text="Load Regex from File",command=lambda: self.load_regex_button_function())
        self.load_regex_button.grid(row=0,column=0,sticky="w",padx=5)
        self.save_regex_button = ttk.Button(self.up_button_frame,text="Save Regex as File",command=lambda: self.save_regex_button_function())
        self.save_regex_button.grid(row=0,column=1,sticky="w",padx=5)
        self.focus_to_last_regex_button = ttk.Button(self.up_button_frame,text="↑",command=lambda: self.focus_to_last_regex_button_function())
        self.focus_to_last_regex_button.grid(row=0,column=2,sticky="w",padx=5)
        self.focus_to_next_regex_button = ttk.Button(self.up_button_frame,text="↓",command=lambda: self.focus_to_next_regex_button_function())
        self.focus_to_next_regex_button.grid(row=0,column=3,sticky="w",padx=5)

        self.show_regex_table_button = ttk.Button(self.up_button_frame,text="Common Regex",command=lambda: self.show_regex_table_function())
        self.show_regex_table_button.grid(row=0,column=5,sticky="e",padx=5)
        self.open_regex_tutorial_button = ttk.Button(self.up_button_frame,text="Regex Tutorial",command=lambda: self.open_regex_tutorial_function())
        self.open_regex_tutorial_button.grid(row=0,column=6,sticky="e",padx=5)

        # 用多行文本框显示多个正则表达式
        # ====== 多行文本框区域 ======
        self.up_text_frame = ttk.Frame(self.up_frame)
        self.up_text_frame.grid(row=2, column=0,sticky="ew",pady=10)
        self.regex_text_scrollbar = ttk.Scrollbar(self.up_text_frame)
        self.regex_text_scrollbar.pack(side="right", fill="y")
        
        self.regex_text_input = tk.Text(self.up_text_frame,wrap="word",yscrollcommand=self.regex_text_scrollbar.set,font=("Consolas", 10),height=10)
        self.regex_text_input.pack(fill="both", expand=True)
        self.regex_text_scrollbar.config(command=self.regex_text_input.yview)
        # 设置捕获组高亮的tag
        self.regex_text_input.tag_configure("capture_group", foreground="#cc6600")
        self.regex_text_input.tag_configure("notcare", background="light gray",foreground="white")
        # 自动高亮捕获组
        #self.regex_text_input.bind("<KeyRelease>", lambda e: self.highlight_capture_groups())
        # 绑定到 Text
        self.regex_text_input.bind("<KeyRelease>", self.update_highlight_items)

        #self.regex_entry_notebook = ttk.Notebook(self.up_text_frame,)# 本身使用默认样式 'TNotebook' :contentReference[oaicite:5]{index=5}


        # 左侧frame ————————————————————————————————————————————————————————————————————————————

        self.left_frame = ttk.Frame(self.main_frame)
        self.left_frame.grid(row=1, column=0, sticky="nsew", padx=5)

        # ====== 标题区域 ======
        self.left_title_frame = ttk.Frame(self.left_frame)
        self.left_title_frame.pack(fill="x", pady=5)
        
        self.left_title_label = ttk.Label(self.left_title_frame, text="Input Text Area", font=("微软雅黑", 11, "bold"))
        self.left_title_label.pack(side="left")
        
        self.left_line_number_label = ttk.Label(self.left_title_frame, text="Current Row: 1/1", font=("Consolas", 10))
        self.left_line_number_label.pack(side="right", padx=(0, 50))
        # ====== 按钮区域 ======
        self.left_btn_frame = ttk.Frame(self.left_frame)
        self.left_btn_frame.pack(fill="x", pady=5)
        self.open_file_button = ttk.Button(self.left_btn_frame,text="Open File",command=lambda: self.open_file_button_function())
        self.open_file_button.grid(row=0,column=0,sticky="w",padx=5)

        style = ttk.Style()
        style.configure("Match.TButton",
                        foreground="#1a5f99")  # 按钮文字颜色
        style.map("Match.TButton",
                  foreground=[("active", "#1a5f99")])# 鼠标悬停颜色

        self.mark_as_mark_button = ttk.Button(self.left_btn_frame,style="Match.TButton",text="Set as Match",command=lambda: self.mark_as_mark_button_function())
        self.mark_as_mark_button.grid(row=0,column=1,sticky="w",padx=5)

        style = ttk.Style()
        style.configure("Target.TButton",
                        foreground="#cc6600")   # 按钮文字颜色
        style.map("Target.TButton",
                  foreground=[("active", "#cc6600")])# 鼠标悬停颜色

        self.mark_as_target_button = ttk.Button(self.left_btn_frame,style="Target.TButton",text="Set as Capture",command=lambda: self.mark_as_target_button_function())
        self.mark_as_target_button.grid(row=0,column=2,sticky="w",padx=5)

        style = ttk.Style()
        style.configure("NotCare.TButton",
                        foreground="#666666")   # 按钮文字颜色
        style.map("Target.TButton",
                  foreground=[("active", "#666666")])# 鼠标悬停颜色

        self.mark_as_not_care_button = ttk.Button(self.left_btn_frame,style="NotCare.TButton",text="Set as Not Care",command=lambda: self.mark_as_not_care_button_function())
        self.mark_as_not_care_button.grid(row=0,column=3,sticky="w",padx=5)
        self.unmark_button = ttk.Button(self.left_btn_frame,text="Unmark",command=lambda: self.unmark_button_function())
        self.unmark_button.grid(row=0,column=4,sticky="w",padx=5)

        # ====== 多行文本框区域 ======
        self.left_text_frame = ttk.Frame(self.left_frame)
        self.left_text_frame.pack(fill="both", expand=True)
        self.left_scrollbar = ttk.Scrollbar(self.left_text_frame)
        self.left_scrollbar.pack(side="right", fill="y")
        self.left_v_scrollbar = ttk.Scrollbar(self.left_text_frame,orient="horizontal")
        self.left_v_scrollbar.pack(side="bottom", fill="x")

        self.left_text_widget = tk.Text(self.left_text_frame,wrap="none",yscrollcommand=self.left_scrollbar.set,xscrollcommand=self.left_v_scrollbar.set,font=("Consolas", 10))
        self.left_text_widget.pack(fill="both", expand=True)
        self.left_scrollbar.config(command=self.left_text_widget.yview)
        self.left_v_scrollbar.config(command=self.left_text_widget.xview)

        # 启用拖放功能
        self.enable_drag_drop()

        self.bind_text_sync_dynamic()
        self.bind_cursor_highlight_sync()
        
        # 绑定 Ctrl+F 快捷键
        self.left_text_widget.bind("<Control-f>", self.show_find_dialog)
        
        # 绑定光标移动事件以更新行号显示
        self.left_text_widget.bind("<KeyRelease>", self.update_line_number_display)

        # 右侧frame ——————————————————————————————————————————————————————————————————————————————————

        self.right_frame =  ttk.Frame(self.main_frame)
        self.right_frame.grid(row=1, column=1, sticky="nsew", padx=5)


        self.right_title_label = ttk.Label(self.right_frame, text="Capture Data Area", font=("微软雅黑", 11, "bold"))
        self.right_title_label.pack(anchor="w", pady=5)

        self.right_btn_frame = ttk.Frame(self.right_frame)
        self.right_btn_frame.pack(fill="x", pady=5)

        self.run_button = ttk.Button(self.right_btn_frame,text="Run",command=lambda: self.run_button_function())
        self.run_button.grid(row=0,column=0,sticky="w",padx=5)
        self.save_as_button = ttk.Button(self.right_btn_frame,text="Save as..",command=lambda: self.save_as_button_function())
        self.save_as_button.grid(row=0,column=3,sticky="w",padx=5)
        self.open_in_excel_button = ttk.Button(self.right_btn_frame,text="Open in Excel",command=lambda: self.open_in_excel_button_function())
        self.open_in_excel_button.grid(row=0,column=4,sticky="w",padx=5)

        self.match_mode_value = tk.IntVar(value=0)
        self.match_mode_checkbutton = ttk.Checkbutton(self.right_btn_frame,text="in Sequence",variable=self.match_mode_value)
        self.match_mode_checkbutton.grid(row=0,column=1,sticky="w",padx=5)

        self.show_row_number_value = tk.IntVar(value=0)
        self.show_row_number_checkbutton = ttk.Checkbutton(self.right_btn_frame,text="Row Num",variable=self.show_row_number_value)
        self.show_row_number_checkbutton.grid(row=0,column=2,sticky="w",padx=5)


        self.right_text_frame = ttk.Frame(self.right_frame)
        self.right_text_frame.pack(fill="both", expand=True)

        self.right_scrollbar = ttk.Scrollbar(self.right_text_frame)
        self.right_scrollbar.pack(side="right", fill="y")

        self.right_v_scrollbar = ttk.Scrollbar(self.right_text_frame,orient="horizontal")
        self.right_v_scrollbar.pack(side="bottom", fill="x")

        self.right_text_widget = tk.Text(self.right_text_frame,wrap="none",yscrollcommand=self.right_scrollbar.set,xscrollcommand=self.right_v_scrollbar.set,font=("Consolas", 10))
        self.right_text_widget.pack(fill="both", expand=True)
        self.right_scrollbar.config(command=self.right_text_widget.yview)
        self.right_v_scrollbar.config(command=self.right_text_widget.xview)
        
        # 绑定双击事件
        self.right_text_widget.bind("<Double-Button-1>", self.on_right_text_double_click)


    def open_file_button_function(self):
        # 打开文件浏览器，选择txt文件
        file_path = filedialog.askopenfilename(
            title="Open TXT File",
            filetypes=[("Text Files", "*.txt"), ("All Files", "*.*")]
        )

        # 如果用户取消选择，直接返回
        if not file_path:
            return

        try:
            # 读取文件内容
            with open(file_path, "r", encoding="utf-8") as f:
                content = f.read()

            # --- 清空已有内容和高亮 ---
            self.left_text_widget.delete("1.0", tk.END)
        
            if hasattr(self, "highlight_items"):
                for item in self.highlight_items:
                    tag = item.get("tag")
                    if tag:
                        try:
                            self.left_text_widget.tag_remove(tag, "1.0", tk.END)
                            self.left_text_widget.tag_delete(tag)
                        except Exception:
                            pass
                self.highlight_items.clear()
            # --- 清空完成 ---
            # 插入新内容
            self.left_text_widget.insert(tk.END, content)

            self.input_string_content = content

        except Exception as e:
            messagebox.showerror("Error", f"Failed to open file:\n{e}")

    def mark_as_mark_button_function(self):
        # 确保 highlight_items 存在
        if not hasattr(self, "highlight_items"):
            self.highlight_items = []

        try:
            sel_start = self.left_text_widget.index("sel.first")
            sel_end   = self.left_text_widget.index("sel.last")
        except:
            messagebox.showerror("Error", "Select text first!")
            return

        selected_text = self.left_text_widget.get(sel_start, sel_end)

        # ------------------------------
        # 不再生成随机 tag，改为统一用 apply_segment_tags
        # ------------------------------
        new_item = {
            "start": sel_start,
            "end": sel_end,
            "text": selected_text,
            "segments": [{
                "text": selected_text,
                "type": "match"
            }],
        }

        # 使用统一 tag 管理函数
        self.apply_segment_tags(new_item)

        # 添加到列表
        self.highlight_items.append(new_item)

        # 生成 regex
        regex_text = self.build_regex_from_segments(new_item)
        new_item["regex"] = regex_text

        self.regex_text_input.delete("1.0", tk.END)
        self.regex_text_input.insert("1.0", regex_text)

        self.left_text_widget.tag_remove("sel", "1.0", "end")
        self.left_text_widget.update_idletasks()

    def build_regex_from_segments(self, item):
        parts = []
        for seg in item["segments"]:
            t = seg["text"]
            tp = seg["type"]

            if tp == "match":
                parts.append(self.translate_to_regex(t))

            elif tp == "capture":
                parts.append("(" + self.generalize_regex_pattern(t) + ")")

            elif tp == "notcare":
                parts.append(r"[\s\S]{0,10000}?")

        return "".join(parts)

    def apply_segment_tags(self, item):
        """
        为 item 的所有 segments 应用 tag。
        如果 item 没有 tag，则自动创建一个新的 tag 基名。
        """

        # 1) 如果没有 tag，就自动生成一个
        if "tag" not in item or not item["tag"]:
            base_tag = f"seg_{id(item)}"
            item["tag"] = base_tag
        else:
            base_tag = item["tag"]

        # 2) 删除旧 tag
        #   注意：不能假设每个 tag_names 都以 base_tag 开头，要安全判断
        for tag in list(self.left_text_widget.tag_names()):
            if tag.startswith(base_tag):
                self.left_text_widget.tag_delete(tag)

        hl_start = item["start"]
        pos = 0
        base_tag = item["tag"]

        for i, seg in enumerate(item["segments"]):
            seg_tag = f"{base_tag}_{i}"

            if seg["type"] == "match":
                self.left_text_widget.tag_configure(seg_tag, background="#cce8ff", foreground="black",selectforeground="gray", selectbackground="white")
            elif seg["type"] == "capture":
                self.left_text_widget.tag_configure(seg_tag, background="#cce8ff", foreground="#cc6600",selectforeground="gray", selectbackground="#cc6600")
            elif seg["type"] == "notcare":
                self.left_text_widget.tag_configure(seg_tag, background="light gray", foreground="white",selectforeground="black", selectbackground="white")

            seg_len = len(seg["text"])
            seg_start = self.left_text_widget.index(f"{hl_start}+{pos}c")
            seg_end   = self.left_text_widget.index(f"{hl_start}+{pos+seg_len}c")

            self.left_text_widget.tag_add(seg_tag, seg_start, seg_end)
            pos += seg_len
        
        self.left_text_widget.update_idletasks()

    def translate_to_regex(self, text: str) -> str:
        """
        将选中的普通文本转换为安全的正则表达式字符串。
        优先级：空白符(泛化) > 转义字符(包含反斜杠) > 数字/字母/中文(保留)
        """
        result = []
        i = 0
        length = len(text)

        while i < length:
            ch = text[i]

            # 1️⃣ 处理连续的空白字符 (空格、Tab、换行)
            if ch.isspace():
                space_count = 0
                while i < length and text[i].isspace():
                    space_count += 1
                    i += 1
                
                # 计算正负 10 的区间
                # 原本 1 个 -> {0, 11}
                # 原本 13 个 -> {3, 23}
                # 注意：下限不能小于 0
                lower_bound = max(0, space_count - 10)
                upper_bound = space_count + 10
                
                # 构建正则表达式，例如 \s{3,23}
                result.append(fr"\s{{{lower_bound},{upper_bound}}}")
                continue

            # 2️⃣ 处理正则表达式中的特殊符号（优先处理反斜杠）
            # 注意：我们将 "\\" 放在首位，确保它是第一个被检测到的特殊字符
            special_chars = "\\.^$*+?{}[]|()"
            if ch in special_chars:
                # 在字符前添加反斜杠进行转义
                result.append("\\" + ch)
                i += 1
                continue

            # 3️⃣ 处理数字、字母和中文
            # 保持原样输出，作为正则中的精确匹配锚点
            if ch.isdigit() or ch.isalpha() or '\u4e00' <= ch <= '\u9fff':
                result.append(ch)
            else:
                # 4️⃣ 对于其他未列出的特殊符号，为了安全也进行转义
                # 或者直接原样输出（如 @, #, % 等在正则中通常不是特殊字符）
                result.append(ch)
            
            i += 1

        return "".join(result)

    # 在input text 文本框绑定修改函数，文本被修改后更新相关变量
    def bind_text_sync_dynamic(self):
        self.left_text_widget.edit_modified(False)
        self.left_text_widget.bind("<<Modified>>", self.on_text_modified_dynamic)

    def on_text_modified_dynamic(self, event=None):
        """
        动态同步 self.highlight_items 逻辑：
        - 如果文本被全部删除或被替换（长度变化大于原长度），则清空 highlight_items
        - 否则保留已有高亮，不更新 start/end/text
        """
        # 清除 Text 的修改标志
        self.left_text_widget.edit_modified(False)

        # 获取当前文本
        current_text = self.left_text_widget.get("1.0", tk.END)

        # 判断是否为整体清空或大幅修改
        if not hasattr(self, "last_text_content"):
            self.last_text_content = current_text

        if len(current_text) == 1 or len(current_text) < len(self.last_text_content) // 2:
            # 整体清空或大幅替换 -> 删除所有 highlight
            if hasattr(self, "highlight_items"):
                for item in self.highlight_items:
                    tag = item.get("tag")
                    if tag:
                        self.left_text_widget.tag_remove(tag, "1.0", tk.END)
                        try:
                            self.left_text_widget.tag_delete(tag)
                        except Exception:
                            pass
                self.highlight_items.clear()

        # 更新 last_text_content
        self.last_text_content = current_text

    def bind_cursor_highlight_sync(self):
        # 鼠标点击移动光标
        self.left_text_widget.bind("<ButtonRelease-1>", self.on_cursor_move_highlight)

    def on_cursor_move_highlight(self, event=None):
        # 获取光标当前索引
        cursor_index = self.left_text_widget.index(tk.INSERT)

        for item in getattr(self, "highlight_items", []):
            start = item.get("start")
            end = item.get("end")

            if start is None or end is None:
                continue  # 如果没有 start/end，就跳过

            # 判断光标是否在该段范围内
            if self.left_text_widget.compare(cursor_index, ">=", start) and \
               self.left_text_widget.compare(cursor_index, "<=", end):
                # 更新 regex_text_input 为该高亮对应 regex
                self.regex_text_input.delete("1.0", tk.END)
                self.regex_text_input.insert("1.0", item.get("regex", ""))
                self.highlight_capture_groups()
                # 同时更新行号显示
                self.update_line_number_display()
                return  # 找到后直接退出
        # 光标不在任何高亮段，则清空 regex_text_input 或保持原样
        # self.regex_text_input.delete(0, tk.END)
        
        # 同时更新行号显示
        self.update_line_number_display()

    def mark_as_target_button_function(self):
        try:
            sel_start = self.left_text_widget.index("sel.first")
            sel_end   = self.left_text_widget.index("sel.last")
        except:
            messagebox.showerror("Error", "Select text first！")
            return

        # 找所属 highlight_item
        item = None
        for it in self.highlight_items:
            if (self.left_text_widget.compare(sel_start, ">=", it["start"]) and
                self.left_text_widget.compare(sel_end, "<=", it["end"])):
                item = it
                break

        if item is None:
            messagebox.showerror("Error", "Select is not within Match section!")
            return

        # 第一次创建 segments
        if "segments" not in item:
            item["segments"] = [{
                "text": item["text"],
                "type": "match"
            }]

        # 1. 切分 segments
        self.split_segments(item, sel_start, sel_end)

        # 2. 给选中的段赋 type="capture"
        for seg in item["segments"]:
            if seg["type"] is None:
                seg["type"] = "capture"

        # 3. 重新生成 regex
        item["regex"] = self.build_regex_from_segments(item)

        # 4. 刷新 UI 高亮
        self.apply_segment_tags(item)

        # 更新右侧文本
        self.regex_text_input.delete("1.0", tk.END)
        self.regex_text_input.insert("1.0", item["regex"])

        self.highlight_capture_groups()

    def mark_as_not_care_button_function(self):
        try:
            sel_start = self.left_text_widget.index("sel.first")
            sel_end   = self.left_text_widget.index("sel.last")
        except Exception:
            messagebox.showerror("Error", "Select text first！")
            return

        # 找 highlight_item
        item = None
        for it in self.highlight_items:
            if (self.left_text_widget.compare(sel_start, ">=", it["start"]) and
                self.left_text_widget.compare(sel_end, "<=", it["end"])):
                item = it
                break

        if item is None:
            messagebox.showerror("Error", "Select is not within Match section!")
            return

        # 确保 segments 存在
        if "segments" not in item or not item["segments"]:
            item["segments"] = [{"text": item["text"], "type": "match"}]

        # 切分段，使选区对齐 segment
        self.split_segments(item, sel_start, sel_end)

        # 计算选区在 highlight_item 内的字符偏移
        try:
            sel_rel_start = self.left_text_widget.count(item['start'], sel_start, "chars")[0]
            sel_rel_end   = self.left_text_widget.count(item['start'], sel_end, "chars")[0]
        except Exception:
            # fallback：选区相对偏移全段
            sel_rel_start = 0
            sel_rel_end = sum(len(s["text"]) for s in item["segments"])

        # 遍历 segments，只修改与选区有重叠的段
        pos = 0
        for seg in item["segments"]:
            seg_start = pos
            seg_end   = pos + len(seg["text"])
            if seg_end > sel_rel_start and seg_start < sel_rel_end:
                seg["type"] = "notcare"
            pos += len(seg["text"])

        # 合并连续相同 type 的段，防止过多 segment
        merged = []
        for s in item["segments"]:
            if merged and merged[-1]["type"] == s["type"]:
                merged[-1]["text"] += s["text"]
            else:
                merged.append({"text": s["text"], "type": s["type"]})
        item["segments"] = merged

        # 重建 regex
        try:
            item["regex"] = self.build_regex_from_segments(item)
        except Exception:
            # fallback 简单拼接
            item["regex"] = "".join(
                ("(.*)" if s["type"]=="notcare" else 
                 "(" + self.generalize_regex_pattern(s["text"]) + ")" if s["type"]=="capture" else 
                 self.translate_to_regex(s["text"]))
                for s in item["segments"]
            )

        # 应用 segment 的 tags，显示灰色
        self.apply_segment_tags(item)

        # 更新右侧 regex 显示
        self.regex_text_input.delete("1.0", tk.END)
        self.regex_text_input.insert("1.0", item.get("regex", ""))
        
        # 高亮 regex 的括号内的 capture group
        self.highlight_capture_groups()

    def unmark_button_function(self):
        cursor = self.left_text_widget.index("insert")

        # 找到 cursor 所属的 highlight_item
        found_item = None
        for item in getattr(self, "highlight_items", []):
            h_start = item.get("start")
            h_end = item.get("end")
            if not h_start or not h_end:
                continue
            try:
                if (self.left_text_widget.compare(cursor, ">=", h_start) and
                    self.left_text_widget.compare(cursor, "<", h_end)):
                    found_item = item
                    break
            except Exception:
                continue

        if not found_item:
            return

        item = found_item

        # --- 1. 计算 cursor 在整个 item 内的 offset ---
        try:
            rel_cursor_off = self.left_text_widget.count(item["start"], cursor, "chars")[0]
        except Exception:
            return

        # --- 2. 找到 cursor 所在 segment ---
        pos = 0
        seg_index = None
        for i, seg in enumerate(item["segments"]):
            seg_len = len(seg["text"])
            if pos <= rel_cursor_off < pos + seg_len:
                seg_index = i
                break
            pos += seg_len

        if seg_index is None:
            return

        seg = item["segments"][seg_index]

        # --- 3. 根据 segment 类型进行处理 ---
        if seg["type"] in ("capture", "notcare"):
            # 这里调用你已有的正式函数
            self.unmark_segment(item, seg)

            # # 更新右侧 regex 显示
            # self.regex_text_input.delete("1.0", tk.END)
            # self.regex_text_input.insert("1.0", item.get("regex", ""))

            # # 高亮 group
            # try:
            #     self.highlight_capture_groups()
            # except Exception:
            #     pass
            return

        # ---------------------------------------------------------
        # 如果是 match segment → 删除整个 highlight（沿用你的旧逻辑）
        # ---------------------------------------------------------
        base_tag = item.get("tag")
        if base_tag:
            for t in list(self.left_text_widget.tag_names()):
                if str(t).startswith(str(base_tag)):
                    try:
                        self.left_text_widget.tag_remove(t, "1.0", "end")
                    except Exception:
                        pass

        # 删除 captures 的 tags
        for cap in item.get("captures", []):
            try:
                if cap.get("tag"):
                    self.left_text_widget.tag_remove(cap["tag"], "1.0", "end")
            except Exception:
                pass

        # 从列表删除
        try:
            self.highlight_items.remove(item)
        except Exception:
            pass

        # 清空 regex 显示
        try:
            self.regex_text_input.delete("1.0", tk.END)
        except Exception:
            pass

    def unmark_segment(self, item, seg_info):
        """安全地删除一个 segment 的 tag，并更新 highlight_item"""
        if not item or not seg_info:
            return

        base_tag = item.get("tag", "")
        # 尝试安全获取 segment tag
        seg_index = item["segments"].index(seg_info)
        c_tag = seg_info.get("tag")
        if not c_tag and base_tag:
            c_tag = f"{base_tag}_{seg_index}"

        # 计算 segment 在 Text 中的起止索引
        try:
            start_index = self.left_text_widget.index(f"{item['start']}+{sum(len(s['text']) for s in item['segments'][:seg_index])}c")
            end_index = self.left_text_widget.index(f"{start_index}+{len(seg_info['text'])}c")
        except Exception:
            start_index = None
            end_index = None

        # 移除该 segment 的 tag
        if c_tag:
            self.remove_tag_safe(c_tag, start=start_index, end=end_index)

        # 如果是 capture 或 notcare，转换为 match
        if seg_info.get("type") in ("capture", "notcare"):
            seg_info["type"] = "match"
            # 清空原 tag 字段（apply_segment_tags 会重新生成）
            seg_info.pop("tag", None)
        else:
            # 如果是 match 段，删除整个 highlight item
            # 移除所有 base_tag 开头的 tag
            if base_tag:
                for t in list(self.left_text_widget.tag_names()):
                    try:
                        if str(t).startswith(base_tag):
                            self.remove_tag_safe(str(t))
                    except Exception:
                        pass
            # 移除所有 captures tag
            for cap in list(item.get("captures", []) or []):
                try:
                    if cap.get("tag"):
                        self.remove_tag_safe(cap.get("tag"), cap.get("start"), cap.get("end"))
                except Exception:
                    pass
            # 从 highlight_items 中删除
            try:
                if item in self.highlight_items:
                    self.highlight_items.remove(item)
            except Exception:
                pass
            return

        # 删除 item 内原 captures 对应 segment 的记录
        try:
            if "captures" in item and isinstance(item["captures"], list):
                for cap in list(item["captures"]):
                    if cap.get("text") == seg_info.get("text"):
                        item["captures"].remove(cap)
        except Exception:
            pass

        # 合并相邻 match 段
        merged = []
        for s in item["segments"]:
            if merged and merged[-1].get("type") == s.get("type") == "match":
                merged[-1]["text"] += s["text"]
            else:
                merged.append({"text": s["text"], "type": s.get("type")})
        item["segments"] = merged

        # 如果 segments 为空，删除整个 highlight
        if not item["segments"]:
            if base_tag:
                for t in list(self.left_text_widget.tag_names()):
                    try:
                        if str(t).startswith(base_tag):
                            self.remove_tag_safe(str(t))
                    except Exception:
                        pass
            try:
                if item in self.highlight_items:
                    self.highlight_items.remove(item)
            except Exception:
                pass
            try:
                self.regex_text_input.delete("1.0", tk.END)
            except Exception:
                pass
            return

        # 更新 item["text"]
        try:
            item["text"] = "".join([s["text"] for s in item["segments"]])
        except Exception:
            pass

        # 更新 regex
        try:
            item["regex"] = self.build_regex_from_segments(item)
        except Exception:
            try:
                item["regex"] = "".join([
                    "(" + self.generalize_regex_pattern(s["text"]) + ")" if s["type"]=="capture" else
                    r"[\s\S]{0,10000}?" if s["type"]=="notcare" else self.translate_to_regex(s["text"])
                    for s in item["segments"]
                ])
            except Exception:
                item["regex"] = item.get("regex", "")

        # 重新应用 tag
        try:
            self.apply_segment_tags(item)
        except Exception:
            pass

        # 更新右侧显示
        try:
            self.regex_text_input.delete("1.0", tk.END)
            self.regex_text_input.insert("1.0", item.get("regex", ""))
            self.highlight_capture_groups()
        except Exception:
            pass

    def remove_tag_safe(self, tag, start=None, end=None):
        """安全移除 tag 并清理样式"""
        try:
            if start and end:
                self.left_text_widget.tag_remove(tag, start, end)
            else:
                ranges = self.left_text_widget.tag_ranges(tag)
                if ranges:
                    for i in range(0, len(ranges), 2):
                        self.left_text_widget.tag_remove(tag, ranges[i], ranges[i+1])
            if not self.left_text_widget.tag_ranges(tag):
                try:
                    self.left_text_widget.tag_delete(tag)
                except Exception:
                    try:
                        self.left_text_widget.tag_config(tag, foreground="", background="")
                    except Exception:
                        pass
        except Exception:
            pass

    def merge_adjacent_match_segments(self, item):
        """
        将 target_item 的主 match tag 覆盖的多个零碎区间合并为一个连续段
        （Text widget 本身可以有多个 range，所以我们手动合并）
        """
        tag = item["tag"]
        ranges = self.left_text_widget.tag_ranges(tag)

        if not ranges:
            return

        # 将 ranges 转成 [(start, end), ...]
        segments = []
        for i in range(0, len(ranges), 2):
            segments.append((ranges[i], ranges[i+1]))

        # 根据 index 排序（避免顺序错乱）
        segments.sort(key=lambda r: float(r[0]))

        # 合并连续区间
        merged = []
        cur_start, cur_end = segments[0]

        for s, e in segments[1:]:
            if self.left_text_widget.compare(s, "<=", cur_end):
                # 有重叠或紧邻 → 合并
                cur_end = e
            else:
                merged.append((cur_start, cur_end))
                cur_start, cur_end = s, e

        merged.append((cur_start, cur_end))

        # 清空原 tag
        self.left_text_widget.tag_remove(tag, "1.0", "end")

        # 添加合并后的区间
        for s, e in merged:
            self.left_text_widget.tag_add(tag, s, e)
            # 更新 item 的 start/end（用第一个区间）
            item["start"], item["end"] = merged[0]

    def rebuild_regex_from_segments(self, item):
        """
        根据当前的匹配区间（含 capture / notcare）重新构建 regex
        """
        start = item["start"]
        end = item["end"]
        hl_text = self.left_text_widget.get(start, end)

        regex = ""
        cur_index = 0

        # 按 capture 排序
        caps = sorted(item.get("captures", []),
                      key=lambda c: float(c["start"]))

        for cap in caps:
            rel_s = self.left_text_widget.count(start, cap["start"], "chars")[0]
            rel_e = self.left_text_widget.count(start, cap["end"], "chars")[0]

            # 普通区域
            normal_text = hl_text[cur_index:rel_s]
            regex += self.translate_to_regex(normal_text)

            # capture / notcare 段
            regex += cap["capture_regex"]

            cur_index = rel_e

        # 最后尾部部分
        tail_text = hl_text[cur_index:]
        regex += self.translate_to_regex(tail_text)

        item["regex"] = regex

    def merge_regex_nodes(self,pattern):
        """
        将正则表达式解析成节点列表，并合并相同类型的连续字符
        节点类型：
            - fixed: 固定字符
            - digit: 数字
            - word: 英文字符
            - chinese: 中文字符
            - space: 空白符
            - special: 其他特殊字符
        返回：
            [{"type": 类型, "content": 内容, "quantifier": 量词}, ...]
        """
        nodes = []
        # 匹配中文、英文、数字、空白、特殊字符
        i = 0
        n = len(pattern)
    
        def get_type(c):
            if c.isdigit():
                return "digit"
            elif re.match(r"[a-zA-Z]", c):
                return "word"
            elif re.match(r"[\u4e00-\u9fff]", c):
                return "chinese"
            elif c.isspace():
                return "space"
            else:
                return "special"

        while i < n:
            c = pattern[i]
            node_type = get_type(c)
            content = c
            quantifier = ""

            # 检查是否有量词在字符后面
            if i + 1 < n and pattern[i + 1] in "*+?":
                quantifier = pattern[i + 1]
                i += 1

            # 合并连续相同类型字符
            j = i + 1
            while j < n:
                c2 = pattern[j]
                t2 = get_type(c2)
                if t2 == node_type:
                    # 检查是否有量词在后面
                    if j + 1 < n and pattern[j + 1] in "*+?":
                        content += c2
                        quantifier += pattern[j + 1]
                        j += 2
                        continue
                    content += c2
                    j += 1
                else:
                    break

            nodes.append({"type": node_type, "content": content, "quantifier": quantifier})
            i = j

        return nodes

    def run_button_function(self):
        """执行正则匹配功能：Mode 0 和 Mode 1 统一使用 finditer 核心逻辑"""
        input_text = self.left_text_widget.get("1.0", "end-1c")
        # 获取高亮项中的正则字符串
        regex_raw_list = [item.get("regex") for item in getattr(self, "highlight_items", []) if item.get("regex")]
        
        if not regex_raw_list:
            messagebox.showerror("Error", "No valid regex rules found.")
            return
        if not input_text.strip():
            messagebox.showerror("Error", "Left text area is empty.")
            return

        self.run_button.config(state="disabled", text="Running...")
        self.update()

        try:
            mode = self.match_mode_value.get()
            show_row = self.show_row_number_value.get()
            self.right_text_widget.delete("1.0", "end")

            # --- 核心搜索阶段：所有模式都先执行全量匹配 ---
            all_regex_matches = [] # 存储结构: [ [ (line_no, val, match_obj), ... ], [ ... ] ]
            
            for regex_str in regex_raw_list:
                try:
                    matches = list(re.finditer(regex_str, input_text))
                except Exception as e:
                    messagebox.showerror("Regex Error", f"Invalid regex:\n{regex_str}\n\n{e}")
                    return

                col_data = []
                current_line_count = 1
                last_line_check_ptr = 0
                
                for m in matches:
                    # 1. 计算增量行号 (核心筛选依据)
                    abs_start = m.start()
                    new_newlines = input_text.count("\n", last_line_check_ptr, abs_start)
                    current_line_count += new_newlines
                    last_line_check_ptr = abs_start
                    
                    # 2. 提取显示内容
                    if m.lastindex:
                        val = " | ".join(str(m.group(i)).replace("\r", "<CR>").replace("\n", "<LF>") for i in range(1, m.lastindex + 1))
                    else:
                        val = m.group(0).replace("\r", "<CR>").replace("\n", "<LF>")
                    
                    # 存储元组：(行号整数, 显示字符串, match对象)
                    col_data.append((current_line_count, val, m))
                
                all_regex_matches.append(col_data)

            # --- 逻辑处理阶段：根据模式筛选结果 ---
            final_columns = [] # 最终显示的列数据
            final_row_numbers = [] # 最终显示的行号数据

            if mode == 0:
                # 模式 0：直接转换
                for col in all_regex_matches:
                    final_columns.append([item[1] for item in col])
                    final_row_numbers.append([str(item[0]) for item in col])
            
            elif mode == 1:
                # 模式 1：瀑布式行号筛选 (Sequence)
                # 筛选逻辑：每一行必须在上一行之后
                filtered_rows = [] # [ [val1, val2...], [val1, val2...] ]
                filtered_rows_no = []
                
                # 以第一个 regex 的每个匹配作为起点
                for i in range(len(all_regex_matches[0])):
                    temp_seq_vals = []
                    temp_seq_nos = []
                    
                    # 第一个元素
                    curr_line_no, curr_val, _ = all_regex_matches[0][i]
                    temp_seq_vals.append(curr_val)
                    temp_seq_nos.append(str(curr_line_no))
                    last_line_no = curr_line_no
                    
                    possible_sequence = True
                    # 从第二个 regex 开始找第一个行号更大的
                    for next_col in all_regex_matches[1:]:
                        found_next = False
                        for next_line_no, next_val, _ in next_col:
                            if next_line_no > last_line_no:
                                temp_seq_vals.append(next_val)
                                temp_seq_nos.append(str(next_line_no))
                                last_line_no = next_line_no
                                found_next = True
                                break
                        
                        if not found_next:
                            possible_sequence = False
                            break
                    
                    if possible_sequence:
                        filtered_rows.append(temp_seq_vals)
                        filtered_rows_no.append(temp_seq_nos)

                # 将筛选后的行数据转回列结构以适配现有的 UI 组装逻辑
                if filtered_rows:
                    num_regex = len(regex_raw_list)
                    for col_idx in range(num_regex):
                        final_columns.append([row[col_idx] for row in filtered_rows])
                        final_row_numbers.append([row[col_idx] for row in filtered_rows_no])

            # --- UI 组装阶段 ---
            if final_columns:
                # 清空位置映射
                self.match_positions = {}
                
                # 1. 表头
                header_parts = []
                for col in final_columns:
                    if show_row == 1: header_parts.append("line")
                    header_parts.append(str(len(col)))
                
                header_line = "\t\t".join(header_parts)
                divider = "-" * 100
                
                # 2. 数据行并记录位置
                max_rows = max(len(col) for col in final_columns)
                result_rows = []
                for r in range(max_rows):
                    row_parts = []
                    for col_idx in range(len(final_columns)):
                        if show_row == 1:
                            row_parts.append(final_row_numbers[col_idx][r] if r < len(final_row_numbers[col_idx]) else "")
                        
                        # 记录位置映射（数据列）
                        if r < len(final_columns[col_idx]):
                            # 计算在输入文本中的位置
                            if mode == 0:
                                # Mode 0: 直接从 all_regex_matches 获取
                                if col_idx < len(all_regex_matches) and r < len(all_regex_matches[col_idx]):
                                    line_no, _, m = all_regex_matches[col_idx][r]
                                    start_pos = f"1.0+{m.start()}c"
                                    end_pos = f"1.0+{m.end()}c"
                                    # 行号从3开始（表头+分隔线+数据）
                                    output_row = r + 3
                                    # 列索引考虑是否显示行号
                                    output_col = col_idx * 2 if show_row == 1 else col_idx
                                    self.match_positions[(output_row, output_col)] = (start_pos, end_pos)
                            elif mode == 1:
                                # Mode 1: 从筛选后的结果获取
                                if col_idx < len(all_regex_matches):
                                    # 找到对应的匹配索引
                                    target_line = int(final_row_numbers[col_idx][r])
                                    for match_idx, (line_no, _, m) in enumerate(all_regex_matches[col_idx]):
                                        if line_no == target_line:
                                            start_pos = f"1.0+{m.start()}c"
                                            end_pos = f"1.0+{m.end()}c"
                                            output_row = r + 3
                                            output_col = col_idx * 2 if show_row == 1 else col_idx
                                            self.match_positions[(output_row, output_col)] = (start_pos, end_pos)
                                            break
                        
                        row_parts.append(final_columns[col_idx][r] if r < len(final_columns[col_idx]) else "")
                    result_rows.append("\t\t".join(row_parts))
                
                final_output = header_line + "\n" + divider + "\n" + "\n".join(result_rows) + "\n"
                self.right_text_widget.insert(tk.END, final_output)

        except Exception as e:
            messagebox.showerror("Runtime Error", f"An error occurred:\n{e}")
        finally:
            self.run_button.config(state="normal", text="Run")

    def highlight_capture_groups(self):
        text = self.regex_text_input.get("1.0", "end-1c")

        # 1. 清除旧的 tag
        self.regex_text_input.tag_remove("capture_group", "1.0", "end")
        self.regex_text_input.tag_remove("notcare", "1.0", "end")
        
        # --- 2. 先计算捕获组括号的范围 (获取所有括号区间的索引) ---
        stack = []
        groups = [] # 存储 (start_index, end_index) 数组
        i = 0
        n = len(text)

        while i < n:
            ch = text[i]
            # 计算反斜杠判定转义
            bs_count = 0
            j = i - 1
            while j >= 0 and text[j] == '\\':
                bs_count += 1
                j -= 1
            is_escaped = (bs_count % 2 == 1)

            if ch == '(' and not is_escaped:
                stack.append(i)
            elif ch == ')' and not is_escaped:
                if stack:
                    start = stack.pop()
                    end = i + 1
                    groups.append((start, end)) # 记录捕获组范围
            i += 1

        # --- 3. 高亮捕获组 (Capture Groups) ---
        for start, end in groups:
            start_index = f"1.0 + {start} chars"
            end_index = f"1.0 + {end} chars"
            self.regex_text_input.tag_add("capture_group", start_index, end_index)

        # --- 4. 高亮 [\s\S]{0,10000}? (通配符)，增加逻辑判断：不能在捕获组内 ---
        targets = [r"[\s\S]{0,10000}?", r".{0,10000}?"] # 建议顺便把刚才讨论的非跨行通配符也加上
        
        for target in targets:
            start_search = 0
            while True:
                idx = text.find(target, start_search)
                if idx == -1:
                    break
                
                # 检查转义
                bs_count = 0
                j = idx - 1
                while j >= 0 and text[j] == '\\':
                    bs_count += 1
                    j -= 1
                
                if bs_count % 2 == 0:  # 未被转义
                    # --- 修改逻辑核心：判断 idx 是否在任何一个 capture_group 范围内 ---
                    is_inside_capture = False
                    for g_start, g_end in groups:
                        # 如果通配符的起始位置在括号区间内
                        if g_start <= idx < g_end:
                            is_inside_capture = True
                            break
                    
                    if not is_inside_capture:
                        start_pos = f"1.0 + {idx} chars"
                        end_pos = f"1.0 + {idx + len(target)} chars"
                        self.regex_text_input.tag_add("notcare", start_pos, end_pos)
                
                start_search = idx + len(target)

    def update_highlight_items(self, event=None):
        """把 regex_text_input 当前行的文本，更新到对应的 highlight_item（基于 left_text_widget 的光标位置判定）"""

        # 1. 获取输入框当前的正则内容
        text = self.regex_text_input.get("1.0", "end-1c").strip()

        # 【新增】校验正则合法性，不合法时输入框变红，但不保存
        if text:
            try:
                re.compile(text)
                self.regex_text_input.config(foreground="black")
            except re.error:
                self.regex_text_input.config(foreground="red")
                self.highlight_capture_groups() # 即使语法错误，也更新capture和not care的高亮
                return # 语法错误时不更新到内存，防止 Run 的时候报错

        # 3. 确定当前光标在左侧文本框的哪个 highlight_item 上
        try:
            cursor = self.left_text_widget.index("insert")
        except Exception:
            cursor = None

        # 确保 highlight_items 存在
        if not hasattr(self, "highlight_items") or self.highlight_items is None:
            self.highlight_items = []

        # 若有有效左侧光标，则找到包含该光标的 item 并更新其 regex
        target_item = None
        if cursor:
            for item in self.highlight_items:
                s = item.get("start")
                e = item.get("end")
                if not s or not e:
                    continue
                try:
                    if (self.left_text_widget.compare(cursor, ">=", s) and
                        self.left_text_widget.compare(cursor, "<=", e)):
                        # 找到对应 item，更新其 regex 字段
                        target_item = item
                        #item["regex"] = text
                        break
                except Exception:
                    # compare 可能因索引格式问题失败，跳过该 item
                    continue

        # 4. 如果找到了对应的 Match 项，更新它的正则
        if target_item:
            target_item["regex"] = text

        # 实时高亮捕获组的括号
        self.highlight_capture_groups()

    def generalize_regex_pattern(self, text):
        t = text.strip()

        # --- 1. 0x 开头的 hex ---
        if re.fullmatch(r"0[xX][A-Fa-f0-9]+", t):
            return r"0x[A-Fa-f0-9]+"

        # --- 2. Hex 字符串 ---
        if re.fullmatch(r"[A-Fa-f0-9]{2,}", t):
            return r"[A-Fa-f0-9]+"

        # --- 3. 数值 (整数、小数、负数) ---
        if re.fullmatch(r"-?\d+(?:\.\d+)?", t):
            return r"-?\d+(?:\.\d+)?"

        # --- 4. IPv4 ---
        if re.fullmatch(r"\d{1,3}(?:\.\d{1,3}){3}", t):
            return r"(?:\d{1,3}\.){3}\d{1,3}"

        # --- 5. MAC 地址 ---
        if re.fullmatch(r"[A-Fa-f0-9]{2}(?:[:-][A-Fa-f0-9]{2}){5}", t):
            return r"[A-Fa-f0-9]{2}(?:[:-][A-Fa-f0-9]{2}){5}"

        # --- 6. 日期 YYYY-MM-DD / YYYY/MM/DD ---
        if re.fullmatch(r"\d{4}[-/]\d{1,2}[-/]\d{1,2}", t):
            return r"\d{4}[-/]\d{1,2}[-/]\d{1,2}"

        # --- 7. 时间 HH:MM[:SS] ---
        if re.fullmatch(r"\d{1,2}:\d{2}(?::\d{2})?", t):
            return r"\d{1,2}:\d{2}(?::\d{2})?"

        # --- 8. 字母数字混合 ---
        if re.fullmatch(r"[A-Za-z0-9]+", t):
            return r"[A-Za-z0-9]+"

        # --- 9. 字母数字空格混合 ---
        if re.fullmatch(r"[A-Za-z0-9\s]+", t):
            return r"[A-Za-z0-9\s\n]*?"

        # --- 9.1 英文数字空格及常用符号 ---
        if re.fullmatch(r"[A-Za-z0-9\s!@#\$%\^&\*\(\)_\+\-=\[\]\{\};:'\",.<>/?\\|`~]+", t):
            return r"[A-Za-z0-9\s!@#\$%\^&\*\(\)_\+\-=\[\]\{\};:'\",.<>/?\\|`~\n]*?"

        # --- 10. 中文 ---
        if re.fullmatch(r"[\u4e00-\u9fa5]+", t):
            return r"[\u4e00-\u9fa5]+"

        # --- 11. 中文 + 英文 ---
        if re.fullmatch(r"[\u4e00-\u9fa5A-Za-z]+", t):
            return r"[\u4e00-\u9fa5A-Za-z]+"

        # --- 12. 中文 + 英文 + 数字 ---
        if re.fullmatch(r"[\u4e00-\u9fa5A-Za-z0-9]+", t):
            return r"[\u4e00-\u9fa5A-Za-z0-9]+"

        # --- 13. 中文 + 英文 + 数字 + 空格 + 中文符号 + 换行 ---
        if re.fullmatch(r"[\u4e00-\u9fa5A-Za-z0-9\s，。！？；：”“‘’（）《》\n]+", t):
            return r"[\u4e00-\u9fa5A-Za-z0-9\s，。！？；：”“‘’（）《》\n]+"

        return self.translate_to_regex(text)

    def split_segments(self, item, abs_sel_start, abs_sel_end):
        """
        将 highlight_item 的 segments 根据选中区间切分成更细的段。
        abs_sel_start / abs_sel_end 是绝对文本坐标（Text widget index）
        """

        hl_start = item["start"]

        # 1. 获取 count 的结果
        res_start = self.left_text_widget.count(hl_start, abs_sel_start, "chars")
        res_end   = self.left_text_widget.count(hl_start, abs_sel_end, "chars")

        # 2. 如果返回值为 None（即无法计算或索引冲突），直接停止运行
        if res_start is None or res_end is None:
            return None

        # 3. 如果结果正常，解包获取相对偏移量
        rel_start = res_start[0]
        rel_end   = res_end[0]

        new_segments = []
        pos = 0

        for seg in item["segments"]:
            t = seg["text"]
            seg_len = len(t)

            if pos + seg_len <= rel_start or pos >= rel_end:
                # 与选区无交集 → 保持原样
                new_segments.append(seg)
            else:
                # 有交集 → 分三段

                # 1) 左边残留
                left_len = max(0, rel_start - pos)
                if left_len > 0:
                    new_segments.append({
                        "text": t[:left_len],
                        "type": seg["type"]
                    })

                # 2) 中间选中部分
                mid_start = max(0, rel_start - pos)
                mid_end   = min(seg_len, rel_end - pos)
                if mid_start < mid_end:
                    new_segments.append({
                        "text": t[mid_start:mid_end],
                        "type": None    # 稍后赋值（capture / notcare）
                    })

                # 3) 右边残留
                right_len = seg_len - mid_end
                if right_len > 0:
                    new_segments.append({
                        "text": t[mid_end:],
                        "type": seg["type"]
                    })

            pos += seg_len

        item["segments"] = new_segments
        return new_segments

    def load_regex_button_function(self):
        # -------------------------------------------------------------
        # 1. 选择文件
        # -------------------------------------------------------------
        file_path = filedialog.askopenfilename(
            title="选择 Regex TXT 文件",
            filetypes=[("Text Files", "*.txt"), ("All Files", "*.*")]
        )

        if not file_path:  # 用户取消
            return

        # -------------------------------------------------------------
        # 2. 读取文件
        # -------------------------------------------------------------
        try:
            with open(file_path, "r", encoding="utf-8") as f:
                lines = [line.strip() for line in f.readlines()]
        except Exception as e:
            messagebox.showerror("Error", f"Coulnd not read file：\n{e}")
            return

        # 去掉空行
        regex_list = [ln for ln in lines if ln]

        if not regex_list:
            messagebox.showwarning("Warning", "File empty, no regex found！")
            return

        # -------------------------------------------------------------
        # 3. 清除当前 highlight_items
        # -------------------------------------------------------------
        self.highlight_items.clear()

        # 记录过程中所有的错误
        error_messages = []

        # 输入文本
        full_text = self.left_text_widget.get("1.0", "end")

        # -------------------------------------------------------------
        # 4. 对每一行 regex 执行同样的逻辑
        # -------------------------------------------------------------
        for regex in regex_list:
            try:
                pattern = re.compile(regex)
            except Exception as e:
                error_messages.append(f"Regex 编译失败: {regex}\n{e}")
                continue

            match = pattern.search(full_text)
            if not match:
                error_messages.append(f"未匹配到内容: {regex}")
                continue

            # ---------------------------------------------------------
            # 4.1 选中第一个匹配内容 → run mark_as_mark_button_function
            # ---------------------------------------------------------
            start_idx = f"1.0+{match.start()}c"
            end_idx   = f"1.0+{match.end()}c"

            try:
                self.left_text_widget.tag_remove("sel", "1.0", "end")
                self.left_text_widget.tag_add("sel", start_idx, end_idx)
                self.mark_as_mark_button_function()
            except Exception as e:
                error_messages.append(f"mark_as_mark_button_function 失败: {regex}\n{e}")
                continue

            # ---------------------------------------------------------
            # 4.2 对其所有 capture group 执行 mark_as_target_button_function
            # ---------------------------------------------------------
            if match.lastindex:  # 有捕获组
                for i in range(1, match.lastindex + 1):
                    try:
                        g_start = match.start(i)
                        g_end   = match.end(i)

                        if g_start == -1 or g_end == -1:
                            continue

                        g_start_idx = f"1.0+{g_start}c"
                        g_end_idx   = f"1.0+{g_end}c"

                        self.left_text_widget.tag_remove("sel", "1.0", "end")
                        self.left_text_widget.tag_add("sel", g_start_idx, g_end_idx)
                        self.mark_as_target_button_function()
                    except Exception as e:
                        error_messages.append(
                            f"mark_as_target_button_function 失败: {regex}\n捕获组 {i}\n{e}"
                        )
                        continue



        # -------------------------------------------------------------
        # 5. 更新 highlight_items 中所有 regex 为 TXT 文件中的内容（带异常处理）
        # -------------------------------------------------------------
        for idx, regex in enumerate(regex_list):
            try:
                if idx < len(self.highlight_items):
                    self.highlight_items[idx]["regex"] = regex
            except Exception as e:
                error_messages.append(f"更新 highlight_item regex 失败，索引 {idx}：{e}")

        # -------------------------------------------------------------
        # 6. 若有错误 → 展示所有错误
        # -------------------------------------------------------------
        if error_messages:
            messagebox.showwarning("Regex Load Finished With Errors", "\n\n".join(error_messages))
        else:
            messagebox.showinfo("Success", "All Regex loaded and applied.")

    def save_regex_button_function(self):
        # -------------------------
        # 1. 检查是否有 regex 可保存
        # -------------------------
        if not hasattr(self, "highlight_items") or len(self.highlight_items) == 0:
            messagebox.showwarning("Warning", "No Regex found！")
            return

        # 收集 regex
        regex_list = []
        for item in self.highlight_items:
            if "regex" in item and item["regex"]:
                regex_list.append(item["regex"])

        if len(regex_list) == 0:
            messagebox.showwarning("Warning", "No Regex found！")
            return

        # -------------------------
        # 2. 弹出保存文件窗口
        # -------------------------
        file_path = filedialog.asksaveasfilename(
            title="保存 Regex 文件",
            defaultextension=".txt",
            filetypes=[("Text Files", "*.txt"), ("All Files", "*.*")]
        )

        if not file_path:  # 用户取消
            return

        # -------------------------
        # 3. 保存到文件
        # -------------------------
        try:
            with open(file_path, "w", encoding="utf-8") as f:
                for regex in regex_list:
                    f.write(regex + "\n")

            messagebox.showinfo("Success", f"Regex successfully save to：\n{file_path}")

        except Exception as e:
            messagebox.showerror("Error", f"File save fail：\n{e}")

    def focus_to_last_regex_button_function(self):
        """跳到上一个 highlight item 的行首（基于 highlight_items 的 start/end）"""
        # 必要性检查
        if not hasattr(self, "highlight_items") or not self.highlight_items:
            messagebox.showinfo("Info", "No highlight items.")
            return

        # 将 start 转为可排序的 tuple (line, col)
        def index_to_tuple(index):
            try:
                s = str(index)
                line, col = s.split(".")
                return int(line), int(col)
            except Exception:
                return (10**9, 10**9)  # 放到最后

        # 收集所有有效 item 的 start/end（跳过没有 start 的）
        items = []
        for it in self.highlight_items:
            s = it.get("start")
            e = it.get("end")
            if not s or not e:
                continue
            items.append((s, e, it))

        if not items:
            messagebox.showinfo("Info", "No valid highlight items with start/end.")
            return

        # 按 start 排序
        items.sort(key=lambda x: index_to_tuple(x[0]))

        # 当前光标位置
        try:
            cur = self.left_text_widget.index("insert")
        except Exception:
            cur = None

        # 如果光标在某个 item 内，则找到该 item 的索引
        cur_idx = None
        if cur:
            for i, (s, e, it) in enumerate(items):
                try:
                    if (self.left_text_widget.compare(cur, ">=", s) and
                        self.left_text_widget.compare(cur, "<=", e)):
                        cur_idx = i
                        break
                except Exception:
                    continue

        # 如果没有在任一 item 内，用 start 与 cur 比较找到前一个 item（start < cur）
        if cur_idx is None and cur:
            prev_i = None
            for i, (s, e, it) in enumerate(items):
                try:
                    if self.left_text_widget.compare(s, "<", cur):
                        prev_i = i
                    else:
                        break
                except Exception:
                    continue
            if prev_i is None:
                messagebox.showinfo("Info", "Already before the first match (no previous match).")
                return
            target_item = items[prev_i]
        else:
            # cur_idx is index of current item; we want previous item
            if cur_idx is None:
                # 没有光标信息，直接取最后一个可用项作为目标
                target_item = items[-1]
            else:
                prev_index = cur_idx - 1
                if prev_index < 0:
                    messagebox.showinfo("Info", "Already at the first match.")
                    return
                target_item = items[prev_index]

        # target_item 是 (start, end, it)
        tgt_start = target_item[0]

        # 移动光标并滚动视图
        try:
            self.left_text_widget.mark_set("insert", tgt_start)
            self.left_text_widget.see(tgt_start)
            # 取消选区，确保视觉正常
            try:
                self.left_text_widget.tag_remove("sel", "1.0", "end")
            except Exception:
                pass
        except Exception:
            messagebox.showerror("Error", "Failed to move cursor to previous match.")

        self.on_cursor_move_highlight()     # 让regex同步显示
        self.left_text_widget.focus_set()   # 关键步骤，光标恢复闪烁


    def focus_to_next_regex_button_function(self):
        """跳到下一个 highlight item 的行首（基于 highlight_items 的 start/end）"""
        if not hasattr(self, "highlight_items") or not self.highlight_items:
            messagebox.showinfo("Info", "No highlight items.")
            return

        def index_to_tuple(index):
            try:
                s = str(index)
                line, col = s.split(".")
                return int(line), int(col)
            except Exception:
                return (10**9, 10**9)

        items = []
        for it in self.highlight_items:
            s = it.get("start")
            e = it.get("end")
            if not s or not e:
                continue
            items.append((s, e, it))

        if not items:
            messagebox.showinfo("Info", "No valid highlight items with start/end.")
            return

        items.sort(key=lambda x: index_to_tuple(x[0]))

        try:
            cur = self.left_text_widget.index("insert")
        except Exception:
            cur = None

        cur_idx = None
        if cur:
            for i, (s, e, it) in enumerate(items):
                try:
                    if (self.left_text_widget.compare(cur, ">=", s) and
                        self.left_text_widget.compare(cur, "<=", e)):
                        cur_idx = i
                        break
                except Exception:
                    continue

        # 如果不在任何 item 内，则找到第一个 start > cur 的 item
        if cur_idx is None and cur:
            next_i = None
            for i, (s, e, it) in enumerate(items):
                try:
                    if self.left_text_widget.compare(s, ">", cur):
                        next_i = i
                        break
                except Exception:
                    continue
            if next_i is None:
                messagebox.showinfo("Info", "Already after the last match (no next match).")
                return
            target_item = items[next_i]
        else:
            # 在某个 item 内，跳到下一个
            if cur_idx is None:
                # 没有光标信息，取第一个
                target_item = items[0]
            else:
                next_index = cur_idx + 1
                if next_index >= len(items):
                    messagebox.showinfo("Info", "Already at the last match.")
                    return
                target_item = items[next_index]

        tgt_start = target_item[0]

        try:
            self.left_text_widget.mark_set("insert", tgt_start)
            self.left_text_widget.see(tgt_start)
            try:
                self.left_text_widget.tag_remove("sel", "1.0", "end")
            except Exception:
                pass
        except Exception:
            messagebox.showerror("Error", "Failed to move cursor to next match.")

        self.on_cursor_move_highlight()     # 让regex同步显示
        self.left_text_widget.focus_set()   # 关键步骤，光标恢复闪烁

    def show_regex_table_function(self):
        """弹出一个窗口，展示常用泛化正则表达式表"""
        # 新建窗口
        win = tk.Toplevel(self)
        win.title("常用泛化正则表达式表")
        win.geometry("1200x420")

        # 文本区域
        text_area = tk.Text(win, wrap="none", font=("Sarasa Mono SC", 12))
        text_area.pack(fill="both", expand=True, padx=10, pady=5)

        # 表格内容
        table_content = r"""NO.     匹配类型                  示例                    泛化正则表达式                                   说明
 1  1个数字                   5                      \d                                               1个0-9的字符
 2  1个或多个数字             31415926               \d+                                              1个或多个数字字符
 3  0个或多个数字             ""、3、31415916        \d*                                              0个或1个或多个数字字符
 4  1个小写字母               j                      [a-z]                                            1个小写字母
 5  1个或多个小写字母         dakbideglskd           [a-z]+                                           1个或多个小写字母
 6  0个或多个小写字母         ""或a、dkciejng        [a-z]*                                           0个或1个或多个小写字母
 7  以0x开头的十六进制数      0x1A3F                 0x[A-Fa-f0-9]+                                   匹配 0x 或 0X 开头的任意长度 hex
 8  十六进制数                1A3F                   [A-Fa-f0-9]+                                     任意长度十六进制字符串
 8  字母和数字                abc123                 [A-Za-z0-9]+                                     仅字母和数字
 9  字母和数字和空格          abc 123                [A-Za-z0-9\s]+                                   允许字母、数字及空格
10  中文                      中文                   [\u4e00-\u9fff]+                                 匹配中文字符
11  中文+英文                 中文abc                [\u4e00-\u9fffA-Za-z]+                           中文和英文混合
12  中文+英文+数字            中文abc123             [\u4e00-\u9fffA-Za-z0-9]+                        中文、英文、数字混合
13  中+文+数+空格+标点        中文 Abc123,。         [\u4e00-\u9fffA-Za-z0-9\s[^\w\s]]+               涵盖几乎所有常见文本字符
14  任意字符                  任意内容               .*?                                              非贪婪匹配任意字符，不包括换行符
15  任意字符                  任意内容               [\s\S]*?                                         非贪婪匹配任意字符，包括换行符
16  任意数值                  123, -45.6             -?\d+(?:\.\d+)?                                  整数、浮点数、正数、负数
17  IPv4地址                  192.168.0.1            (?:\d{1,3}\.){3}\d{1,3}                          典型 IPv4 地址
18  MAC地址                   01:23:45:67:89:AB      [A-Fa-f0-9]{2}(?:[:-][A-Fa-f0-9]{2}){5}          MAC 地址，可用冒号或短横分隔
19  日期                      2025-12-09             \d{4}[-/]\d{1,2}[-/]\d{1,2}                      YYYY-MM-DD 或 YYYY/MM/DD
20  时间                      14:30 或 14:30:59      \d{1,2}:\d{2}(?::\d{2})?                         HH:MM 或 HH:MM:SS
21  空格                      空格                   " " 或 [ ] 或 \x20                               1个空格
22  空白符                    空格 \t \r \n          \s                                               1个空格、制表符\t、换行\n、回车\r
"""

        text_area.insert("1.0", table_content)
        #text_area.config(state="disabled")  # 设置只读

    def open_regex_tutorial_function(self):
        """直接打开默认浏览器，跳转到 Python 正则表达式教程"""
        try:
            webbrowser.open("https://deerchao.cn/tutorials/regex/regex.htm?utm_source=chatgpt.com")
        except Exception as e:
            tk.messagebox.showerror("Error", f"Could not open browser:\n{e}")

    def save_as_button_function(self):
        try:
            # 1. 获取文本内容
            content = self.right_text_widget.get("1.0", "end-1c")
            
            if not content.strip():
                messagebox.showwarning("Warning", "Capture Data Area is empty!")
                return

            # 2. 选择保存位置和格式
            file_path = filedialog.asksaveasfilename(
                title="Save As",
                defaultextension=".txt",
                filetypes=[("Text Files", "*.txt"), ("Excel Files", "*.xlsx"), ("All Files", "*.*")]
            )

            # 用户取消
            if not file_path:
                return
            
            # 3. 根据文件扩展名保存
            if file_path.endswith('.xlsx'):
                # 保存为 Excel
                if not OPENPYXL_AVAILABLE:
                    messagebox.showerror("Error", "openpyxl library not found.\nPlease install it: pip install openpyxl")
                    return
                
                wb = openpyxl.Workbook()
                ws = wb.active
                
                lines = content.split('\n')
                for row_idx, line in enumerate(lines, start=1):
                    if line.strip():
                        cells = line.split('\t\t')
                        for col_idx, cell_value in enumerate(cells, start=1):
                            ws.cell(row=row_idx, column=col_idx, value=cell_value)
                
                wb.save(file_path)
                messagebox.showinfo("Saved", f"File saved to:\n{file_path}")
            else:
                # 保存为文本文件
                with open(file_path, "w", encoding="utf-8") as f:
                    f.write(content)
                messagebox.showinfo("Saved", f"File saved to:\n{file_path}")

        except Exception as e:
            messagebox.showerror("Error", f"Failed to save file:\n{e}")

    def open_in_excel_button_function(self):
        if not OPENPYXL_AVAILABLE:
            messagebox.showerror("Error", "openpyxl library not found.\nPlease install it: pip install openpyxl")
            return
        
        try:
            # 获取右侧文本框内容
            content = self.right_text_widget.get("1.0", "end-1c")
            
            if not content.strip():
                messagebox.showwarning("Warning", "Capture Data Area is empty!")
                return
            
            # 创建新的 Excel 工作簿
            wb = openpyxl.Workbook()
            ws = wb.active
            
            # 按行分割内容
            lines = content.split('\n')
            
            # 写入 Excel
            for row_idx, line in enumerate(lines, start=1):
                if line.strip():  # 跳过空行
                    # 按制表符分割列
                    cells = line.split('\t\t')
                    for col_idx, cell_value in enumerate(cells, start=1):
                        ws.cell(row=row_idx, column=col_idx, value=cell_value)
            
            # 创建临时文件
            temp_file = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
            temp_path = temp_file.name
            temp_file.close()
            
            # 保存 Excel 文件
            wb.save(temp_path)
            
            # 用系统默认程序打开
            os.startfile(temp_path)
            
        except Exception as e:
            messagebox.showerror("Error", f"Failed to open in Excel:\n{e}")

    def show_about_window(self):
        top = tk.Toplevel(self)
        top.title("Coffee!")
        # 设置固定窗口大小
        win_width = 300
        win_height = 300

        # 获取主窗口的位置和尺寸
        self.update_idletasks()
        root_x = self.winfo_x()
        root_y = self.winfo_y()
        root_width = self.winfo_width()
        root_height = self.winfo_height()

        # 计算居中位置
        pos_x = root_x + (root_width - win_width) // 2
        pos_y = root_y + (root_height - win_height) // 2

        # 设置 Toplevel 窗口的大小和位置
        top.geometry(f"{win_width}x{win_height}+{pos_x}+{pos_y}")
    
        # 加载大图
        try:
            # 解码 base64 数据
            img_data = base64.b64decode(img_base64)
            img_pil = Image.open(BytesIO(img_data))
            img_resized = img_pil.resize((200, 200), Image.Resampling.LANCZOS)
            photo = ImageTk.PhotoImage(img_resized)

            # 显示图片
            label = tk.Label(top, image=photo)
            label.image = photo  # 防止被垃圾回收
            label.pack(pady=10)
        except Exception as e:
            tk.Label(top, text=f"图片显示失败: {e}").pack()

        tk.Label(top, text="If this tool saved your time,").pack()
        tk.Label(top, text="feel free to buy me a coffee.").pack()
        tk.Label(top, text="Your support means a lot to me!").pack()

    def enable_drag_drop(self):
        """启用拖放文件功能"""
        # Windows 系统支持 DND
        try:
            self.left_text_widget.drop_target_register('DND_Files')
            self.left_text_widget.dnd_bind('<<Drop>>', self.on_drop)
        except:
            # 如果上面的方法不可用，使用 tkinterdnd2 方式
            if TKINTERDND2_AVAILABLE:
                self.left_text_widget.drop_target_register(DND_FILES)
                self.left_text_widget.dnd_bind('<<Drop>>', self.on_drop)
            else:
                # 如果没有 tkinterdnd2，使用原生 tkinter 方法
                self.left_text_widget.bind('<Button-1>', self.on_text_click)

    def on_drop(self, event):
        """处理拖放事件"""
        try:
            # 获取拖放的文件路径
            files = event.data
            
            # 处理文件路径（可能包含花括号或多个文件）
            if isinstance(files, str):
                # 移除花括号
                files = files.strip('{}').strip()
                # 分割多个文件（如果有）
                file_list = files.split('} {')
                file_path = file_list[0].strip()
            else:
                file_path = str(files)
            
            # 读取文件内容
            with open(file_path, 'r', encoding='utf-8') as f:
                content = f.read()
            
            # 清空已有内容和高亮
            self.left_text_widget.delete("1.0", tk.END)
            
            if hasattr(self, "highlight_items"):
                for item in self.highlight_items:
                    tag = item.get("tag")
                    if tag:
                        try:
                            self.left_text_widget.tag_remove(tag, "1.0", tk.END)
                            self.left_text_widget.tag_delete(tag)
                        except Exception:
                            pass
                self.highlight_items.clear()
            
            # 插入新内容
            self.left_text_widget.insert(tk.END, content)
            self.input_string_content = content
            
        except Exception as e:
            messagebox.showerror("Error", f"Failed to load file:\n{e}")
        
        return 'break'

    def on_text_click(self, event):
        """备用方法：处理点击事件（如果拖放不可用）"""
        pass

    def show_find_dialog(self, event=None):
        """显示查找对话框"""
        if self.find_window and self.find_window.winfo_exists():
            self.find_window.lift()
            self.find_window.focus()
            return "break"
        
        self.find_window = tk.Toplevel(self)
        self.find_window.title("Find")
        self.find_window.geometry("450x120")
        self.find_window.resizable(False, False)
        
        self.find_window.transient(self)
        # 移除 grab_set，允许切换回主窗口
        
        ttk.Label(self.find_window, text="Find:").grid(row=0, column=0, padx=10, pady=10, sticky="w")
        
        self.search_combobox = ttk.Combobox(self.find_window, values=self.search_history, width=40)
        self.search_combobox.grid(row=0, column=1, padx=10, pady=10, columnspan=3)
        self.search_combobox.focus()
        
        self.search_combobox.bind("<Return>", lambda e: self.find_next())
        
        # Find Previous 在左侧
        find_prev_btn = ttk.Button(self.find_window, text="Find Previous", command=self.find_previous, width=15)
        find_prev_btn.grid(row=1, column=1, padx=5, pady=5, sticky="w")
        
        # Find Next 在中间
        find_next_btn = ttk.Button(self.find_window, text="Find Next", command=self.find_next, width=15)
        find_next_btn.grid(row=1, column=2, padx=5, pady=5)
        
        # Count 按钮在右侧
        count_btn = ttk.Button(self.find_window, text="Count", command=self.count_matches, width=10)
        count_btn.grid(row=1, column=3, padx=5, pady=5, sticky="e")
        
        # 底部状态标签（预留位置）
        self.find_status_label = ttk.Label(self.find_window, text="", foreground="blue")
        self.find_status_label.grid(row=2, column=1, columnspan=3, padx=10, pady=5, sticky="w")
        
        return "break"

    def find_next(self):
        """从当前光标位置向后查找"""
        search_text = self.search_combobox.get().strip()
        
        if not search_text:
            messagebox.showwarning("Find", "Please enter search text")
            return
        
        self._update_search_history(search_text)
        
        current_pos = self.left_text_widget.index(tk.INSERT)
        
        pos = self.left_text_widget.search(search_text, current_pos, tk.END, nocase=True)
        
        if pos:
            end_pos = f"{pos}+{len(search_text)}c"
            self.left_text_widget.mark_set(tk.INSERT, end_pos)
            
            # 选中找到的文本
            self.left_text_widget.tag_remove("sel", "1.0", tk.END)
            self.left_text_widget.tag_add("sel", pos, end_pos)
            
            # 计算行号
            line_num = int(pos.split('.')[0])
            total_lines = int(self.left_text_widget.index('end-1c').split('.')[0])
            
            # 更新状态显示
            if hasattr(self, 'find_status_label'):
                self.find_status_label.config(text=f"Line {line_num}/{total_lines}")
            
            self.left_text_widget.see(end_pos)
            self.left_text_widget.focus()
        else:
            pos = self.left_text_widget.search(search_text, "1.0", current_pos, nocase=True)
            if pos:
                end_pos = f"{pos}+{len(search_text)}c"
                self.left_text_widget.mark_set(tk.INSERT, end_pos)
                self.left_text_widget.tag_remove("sel", "1.0", tk.END)
                self.left_text_widget.tag_add("sel", pos, end_pos)
                
                line_num = int(pos.split('.')[0])
                total_lines = int(self.left_text_widget.index('end-1c').split('.')[0])
                
                if hasattr(self, 'find_status_label'):
                    self.find_status_label.config(text=f"Line {line_num}/{total_lines} (wrapped)")
                
                self.left_text_widget.see(end_pos)
                self.left_text_widget.focus()
            else:
                if hasattr(self, 'find_status_label'):
                    self.find_status_label.config(text=f"'{search_text}' not found")

    def find_previous(self):
        """从当前光标位置向前查找"""
        search_text = self.search_combobox.get().strip()
        
        if not search_text:
            messagebox.showwarning("Find", "Please enter search text")
            return
        
        self._update_search_history(search_text)
        
        current_pos = self.left_text_widget.index(tk.INSERT)
        
        pos = self.left_text_widget.search(search_text, current_pos, "1.0", backwards=True, nocase=True)
        
        if pos:
            end_pos = f"{pos}+{len(search_text)}c"
            self.left_text_widget.mark_set(tk.INSERT, pos)
            
            # 选中找到的文本
            self.left_text_widget.tag_remove("sel", "1.0", tk.END)
            self.left_text_widget.tag_add("sel", pos, end_pos)
            
            # 计算行号
            line_num = int(pos.split('.')[0])
            total_lines = int(self.left_text_widget.index('end-1c').split('.')[0])
            
            # 更新状态显示
            if hasattr(self, 'find_status_label'):
                self.find_status_label.config(text=f"Line {line_num}/{total_lines}")
            
            self.left_text_widget.see(pos)
            self.left_text_widget.focus()
        else:
            pos = self.left_text_widget.search(search_text, tk.END, current_pos, backwards=True, nocase=True)
            if pos:
                end_pos = f"{pos}+{len(search_text)}c"
                self.left_text_widget.mark_set(tk.INSERT, pos)
                self.left_text_widget.tag_remove("sel", "1.0", tk.END)
                self.left_text_widget.tag_add("sel", pos, end_pos)
                
                line_num = int(pos.split('.')[0])
                total_lines = int(self.left_text_widget.index('end-1c').split('.')[0])
                
                if hasattr(self, 'find_status_label'):
                    self.find_status_label.config(text=f"Line {line_num}/{total_lines} (wrapped)")
                
                self.left_text_widget.see(pos)
                self.left_text_widget.focus()
            else:
                if hasattr(self, 'find_status_label'):
                    self.find_status_label.config(text=f"'{search_text}' not found")

    def _update_search_history(self, search_text):
        """更新搜索历史记录"""
        if search_text in self.search_history:
            self.search_history.remove(search_text)
        
        self.search_history.insert(0, search_text)
        
        if len(self.search_history) > 10:
            self.search_history = self.search_history[:10]
        
        if hasattr(self, 'search_combobox'):
            self.search_combobox['values'] = self.search_history

    def count_matches(self):
        """统计匹配次数"""
        search_text = self.search_combobox.get().strip()
        
        if not search_text:
            messagebox.showwarning("Find", "Please enter search text")
            return
        
        self._update_search_history(search_text)
        
        # 获取所有文本
        all_text = self.left_text_widget.get("1.0", tk.END)
        
        # 统计出现次数（大小写不敏感）
        count = 0
        start_pos = "1.0"
        while True:
            pos = self.left_text_widget.search(search_text, start_pos, tk.END, nocase=True)
            if not pos:
                break
            count += 1
            start_pos = f"{pos}+1c"
        
        # 显示结果
        if hasattr(self, 'find_status_label'):
            self.find_status_label.config(text=f"Found {count} occurrence(s) of '{search_text}'")
    
    def update_line_number_display(self, event=None):
        """更新左侧文本框的行号显示"""
        try:
            # 获取当前光标位置
            cursor_pos = self.left_text_widget.index(tk.INSERT)
            current_line = int(cursor_pos.split('.')[0])
            
            # 获取总行数
            total_lines = int(self.left_text_widget.index('end-1c').split('.')[0])
            
            # 更新标签
            self.left_line_number_label.config(text=f"Current Row: {current_line}/{total_lines}")
        except Exception:
            pass
    
    def on_right_text_double_click(self, event):
        """处理右侧文本框的双击事件，定位到左侧对应位置"""
        try:
            # 获取双击位置
            index = self.right_text_widget.index(f"@{event.x},{event.y}")
            row, col_char = map(int, index.split('.'))
            
            # 跳过表头和分隔线
            if row < 3:
                return
            
            # 计算列索引（通过统计制表符数量）
            line_text = self.right_text_widget.get(f"{row}.0", f"{row}.{col_char}")
            col_index = line_text.count('\t\t')
            
            # 查找映射
            if (row, col_index) in self.match_positions:
                start_pos, end_pos = self.match_positions[(row, col_index)]
                
                # 选中左侧文本
                self.left_text_widget.tag_remove("sel", "1.0", tk.END)
                self.left_text_widget.tag_add("sel", start_pos, end_pos)
                
                # 移动光标并滚动到可见位置
                self.left_text_widget.mark_set(tk.INSERT, start_pos)
                self.left_text_widget.see(start_pos)
                self.left_text_widget.focus_set()
                
                # 更新行号显示
                self.update_line_number_display()
        except Exception as e:
            pass  # 静默失败

if __name__ == "__main__":
    app = MainGUI()
    app.mainloop()

